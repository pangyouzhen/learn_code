import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
import pandas as pd

def write_dataframes_to_excel(filename, *dataframes, titles=None, sheet_name=None):
    """
    将多个 DataFrame 写入同一个 sheet 中，每个 DataFrame 用边框框住，并可加标题。
    每两个 DataFrame 之间空一行。

    Args:
        filename (str): Excel 文件名。
        *dataframes (pd.DataFrame): 任意数量的 pandas DataFrame。
        titles (list of str, optional): 每个 DataFrame 的标题。可选。
        sheet_name (str, optional): 指定写入的 sheet 名称。默认使用 active sheet。
    """
    workbook = openpyxl.Workbook()

    # 删除默认创建的 Sheet，如果指定了新 sheet_name
    if sheet_name:
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet = workbook.active
        if sheet_name:
            sheet.title = sheet_name

    current_row = 1

    # 样式定义
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    for i, df in enumerate(dataframes):
        # 写标题
        if titles and i < len(titles) and titles[i]:
            title = titles[i]
            end_col = len(df.columns)
            title_cell = sheet.cell(row=current_row, column=1, value=title)
            title_cell.font = bold_font
            title_cell.alignment = center_align
            if end_col > 1:
                sheet.merge_cells(start_row=current_row, start_column=1,
                                  end_row=current_row, end_column=end_col)
            current_row += 1

        start_row = current_row

        # 写列名
        for col_index, column in enumerate(df.columns):
            cell = sheet.cell(row=current_row, column=col_index + 1, value=column)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border

        # 写数据
        for row_index, row_data in df.iterrows():
            current_row += 1
            for col_index, value in enumerate(row_data):
                cell = sheet.cell(row=current_row, column=col_index + 1, value=value)
                cell.border = border

        # 加边框
        end_row = current_row
        end_col = len(df.columns)
        for row in range(start_row, end_row + 1):
            for col in range(1, end_col + 1):
                sheet.cell(row=row, column=col).border = border

        current_row += 2  # 空一行再接下一个 DataFrame

    workbook.save(filename)

if __name__ == '__main__':
    df1 = pd.DataFrame({'col1': [1, 2], 'col2': [3, 4]})
    df2 = pd.DataFrame({'A': [5, 6, 7], 'B': [8, 9, 10], 'C': [11, 12, 13]})
    df3 = pd.DataFrame({'X': [14], 'Y': [15]})

    titles = ['表一：示例数据1', '表二：示例数据2', '表三：示例数据3']
    excel_filename = 'output_with_sheetname.xlsx'
    sheet_name = '结果统计'

    write_dataframes_to_excel(excel_filename, df1, df2, df3, titles=titles, sheet_name=sheet_name)

    print(f"数据已写入到 {excel_filename} 的 sheet: {sheet_name}")
